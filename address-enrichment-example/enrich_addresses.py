#!/usr/bin/env python3
"""
Address enrichment pipeline for cold-calling prep.

Works in two modes:
  1. Heuristic-only mode, which classifies obvious government/school/etc. rows.
  2. Google Places mode, enabled by GOOGLE_PLACES_API_KEY or --provider google.

CSV files work with Python's standard library. Excel files require openpyxl.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any


DEFAULT_FIELD_MASK = (
    "places.id,"
    "places.displayName,"
    "places.formattedAddress,"
    "places.nationalPhoneNumber,"
    "places.internationalPhoneNumber,"
    "places.websiteUri,"
    "places.googleMapsUri,"
    "places.businessStatus,"
    "places.types,"
    "places.primaryType,"
    "places.rating,"
    "places.userRatingCount"
)

FIELD_PRESETS = {
    "ids": "places.id,places.name",
    "basic": (
        "places.id,"
        "places.displayName,"
        "places.formattedAddress,"
        "places.googleMapsUri,"
        "places.businessStatus,"
        "places.primaryType,"
        "places.types"
    ),
    "contact": DEFAULT_FIELD_MASK,
}

OUTPUT_COLUMNS = [
    "business_name",
    "normalized_address",
    "category",
    "do_not_call",
    "exclude_reason",
    "phone",
    "international_phone",
    "website",
    "google_maps_url",
    "google_place_id",
    "google_business_status",
    "google_primary_type",
    "google_types",
    "rating",
    "rating_count",
    "confidence",
    "enrichment_status",
    "source_provider",
    "source_url",
    "notes",
    "address_match",
    "query_used",
]

GOVERNMENT_KEYWORDS = [
    "city hall",
    "courthouse",
    "county",
    "federal",
    "government",
    "municipal",
    "state capitol",
    "public works",
    "police department",
    "fire department",
]

SCHOOL_KEYWORDS = [
    "school",
    "university",
    "college",
    "academy",
    "school district",
]

HEALTHCARE_KEYWORDS = [
    "hospital",
    "clinic",
    "medical",
    "health",
    "urgent care",
]

LANDMARK_KEYWORDS = [
    "landmark",
    "museum",
    "monument",
    "tourist attraction",
    "park",
    "stadium",
    "non-business",
]

INVALID_KEYWORDS = [
    "invalid",
    "test failure",
    "bikini bottom",
]

BUSINESS_KEYWORDS = [
    "business",
    "commercial",
    "corporate",
    "headquarters",
    "hq",
    "office building",
    "studio",
]

GOVERNMENT_TYPES = {
    "city_hall",
    "courthouse",
    "embassy",
    "fire_station",
    "local_government_office",
    "police",
    "post_office",
}

SCHOOL_TYPES = {
    "primary_school",
    "school",
    "secondary_school",
    "university",
}

HEALTHCARE_TYPES = {
    "dentist",
    "doctor",
    "drugstore",
    "hospital",
    "pharmacy",
    "physiotherapist",
}

LANDMARK_TYPES = {
    "amusement_park",
    "aquarium",
    "art_gallery",
    "cemetery",
    "church",
    "hindu_temple",
    "library",
    "museum",
    "mosque",
    "park",
    "stadium",
    "synagogue",
    "tourist_attraction",
    "zoo",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Enrich address CSV rows with business/contact details and exclusion flags."
    )
    parser.add_argument("--input", required=True, help="Input CSV path.")
    parser.add_argument("--output", required=True, help="Output enriched CSV path.")
    parser.add_argument(
        "--provider",
        choices=["auto", "heuristic", "google"],
        default="auto",
        help="auto uses Google Places when GOOGLE_PLACES_API_KEY is set, otherwise heuristics only.",
    )
    parser.add_argument(
        "--cache",
        default=".cache/places_cache.json",
        help="JSON cache path for API responses.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Optional max number of input rows to process.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Excel sheet name to read. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--max-results",
        type=int,
        default=1,
        help="Maximum Google Places matches to write per input address. Use 3-5 for multi-tenant buildings.",
    )
    parser.add_argument(
        "--search-strategy",
        choices=["exact", "expanded"],
        default="exact",
        help="exact searches the address once. expanded also searches business-at-address variants for multi-tenant buildings.",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.1,
        help="Seconds to sleep between uncached Google requests.",
    )
    parser.add_argument(
        "--field-preset",
        choices=sorted(FIELD_PRESETS),
        default="contact",
        help="Google Places field set. Use basic to diagnose key/API setup before requesting phone/website fields.",
    )
    parser.add_argument(
        "--field-mask",
        default=None,
        help="Custom Google Places field mask. Overrides --field-preset.",
    )
    return parser.parse_args()


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(cache, handle, indent=2, sort_keys=True)


def compact(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def build_query(row: dict[str, str]) -> str:
    parts = [
        compact(row.get("address")),
        compact(row.get("city")),
        compact(row.get("state")),
        compact(row.get("postal_code")),
    ]
    return ", ".join(part for part in parts if part)


def build_search_queries(row: dict[str, str], strategy: str) -> list[str]:
    exact = build_query(row)
    if strategy == "exact":
        return [exact]

    street = compact(row.get("address"))
    city = compact(row.get("city"))
    state = compact(row.get("state"))
    postal_code = compact(row.get("postal_code"))
    locality = " ".join(part for part in [city, state, postal_code] if part)
    compact_address = " ".join(part for part in [street, locality] if part)

    candidates = [
        exact,
        f"business at {compact_address}",
        f"businesses at {compact_address}",
        f"{compact_address} business",
        f"{compact_address} tenants",
    ]
    deduped: list[str] = []
    for candidate in candidates:
        if candidate and candidate not in deduped:
            deduped.append(candidate)
    return deduped


def normalized_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def has_any_keyword(text: str, keywords: list[str]) -> bool:
    folded = f" {normalized_text(text)} "
    return any(f" {normalized_text(keyword)} " in folded for keyword in keywords)


def display_name(place: dict[str, Any] | None) -> str:
    if not place:
        return ""
    name = place.get("displayName")
    if isinstance(name, dict):
        return compact(name.get("text"))
    return compact(name)


def address_match(row: dict[str, str], place: dict[str, Any] | None) -> str:
    if not place:
        return ""
    input_street = normalized_text(compact(row.get("address")))
    formatted = normalized_text(compact(place.get("formattedAddress")))
    postal_code = normalized_text(compact(row.get("postal_code")))
    if input_street and input_street in formatted:
        if not postal_code or postal_code in formatted:
            return "exact"
        return "street_match_zip_differs"
    input_number = input_street.split(" ", 1)[0] if input_street else ""
    if input_number and f" {input_number} " in f" {formatted} ":
        return "possible"
    return "nearby_or_related"


def choose_category(
    row: dict[str, str],
    place: dict[str, Any] | None,
    status: str,
) -> tuple[str, bool, str, str]:
    text_parts = [
        build_query(row),
        compact(row.get("source_notes")),
        display_name(place),
        compact(place.get("formattedAddress") if place else ""),
        compact(place.get("primaryType") if place else ""),
        " ".join(place.get("types", []) if place else []),
    ]
    text = " ".join(text_parts)
    types = set(place.get("types", []) if place else [])

    if status == "no_match" or has_any_keyword(text, INVALID_KEYWORDS):
        return "invalid_or_unmatched", True, "Invalid address or no confident place match.", "0.20"
    if types & GOVERNMENT_TYPES or has_any_keyword(text, GOVERNMENT_KEYWORDS):
        return "government", True, "Government or public-sector building.", "0.90"
    if types & SCHOOL_TYPES or has_any_keyword(text, SCHOOL_KEYWORDS):
        return "school", True, "School, college, or university.", "0.90"
    if types & HEALTHCARE_TYPES or has_any_keyword(text, HEALTHCARE_KEYWORDS):
        return "healthcare", False, "", "0.85"
    if types & LANDMARK_TYPES or has_any_keyword(text, LANDMARK_KEYWORDS):
        return "landmark_or_non_business", True, "Landmark, museum, park, or other non-cold-call target.", "0.85"
    if has_any_keyword(text, BUSINESS_KEYWORDS):
        return "business", False, "", "0.65"
    if place:
        return "business", False, "", "0.75"
    return "ambiguous", False, "", "0.35"


class PlacesApiError(Exception):
    pass


def google_places_search(
    query: str,
    api_key: str,
    field_mask: str,
    max_results: int,
) -> dict[str, Any]:
    request = urllib.request.Request(
        "https://places.googleapis.com/v1/places:searchText",
        data=json.dumps(
            {
                "textQuery": query,
                "regionCode": "US",
                "pageSize": max(1, min(max_results, 20)),
            }
        ).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": field_mask,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        body = error.read().decode("utf-8", errors="replace")
        raise PlacesApiError(f"HTTP {error.code} {error.reason}: {body}") from error


def lookup_places(
    query: str,
    provider: str,
    api_key: str | None,
    field_mask: str,
    cache: dict[str, Any],
    sleep_seconds: float,
    max_results: int,
) -> tuple[list[dict[str, Any]], str, str]:
    if provider == "heuristic":
        return [], "heuristic_only", ""

    cache_key = f"google:{max_results}:{query}"
    if cache_key in cache:
        cached = cache[cache_key]
        if cached.get("status") != "api_error":
            if "places" in cached:
                return cached.get("places") or [], cached.get("status", "cached"), cached.get("error", "")
            place = cached.get("place")
            return ([place] if place else []), cached.get("status", "cached"), cached.get("error", "")

    try:
        payload = google_places_search(query, api_key or "", field_mask, max_results)
        places = payload.get("places") or []
        status = "enriched" if places else "no_match"
        cache[cache_key] = {"status": status, "places": places}
        time.sleep(sleep_seconds)
        return places, status, ""
    except (PlacesApiError, urllib.error.URLError, TimeoutError) as error:
        message = str(error)
        return [], "api_error", message


def enriched_row(
    row: dict[str, str],
    place: dict[str, Any] | None,
    provider: str,
    status: str,
    error: str,
    match_rank: int = 0,
) -> dict[str, str]:
    category, do_not_call, exclude_reason, confidence = choose_category(row, place, status)
    types = place.get("types", []) if place else []
    google_maps_url = compact(place.get("googleMapsUri") if place else "")

    output = dict(row)
    output.update(
        {
            "business_name": display_name(place),
            "normalized_address": compact(place.get("formattedAddress") if place else ""),
            "category": category,
            "do_not_call": "TRUE" if do_not_call else "FALSE",
            "exclude_reason": exclude_reason,
            "phone": compact(place.get("nationalPhoneNumber") if place else ""),
            "international_phone": compact(place.get("internationalPhoneNumber") if place else ""),
            "website": compact(place.get("websiteUri") if place else ""),
            "google_maps_url": google_maps_url,
            "google_place_id": compact(place.get("id") if place else ""),
            "google_business_status": compact(place.get("businessStatus") if place else ""),
            "google_primary_type": compact(place.get("primaryType") if place else ""),
            "google_types": "|".join(types),
            "rating": compact(place.get("rating") if place else ""),
            "rating_count": compact(place.get("userRatingCount") if place else ""),
            "confidence": confidence,
            "enrichment_status": status,
            "source_provider": provider,
            "source_url": google_maps_url,
            "notes": error,
            "match_rank": str(match_rank) if match_rank else "",
            "address_match": address_match(row, place),
            "query_used": compact(place.get("_query_used") if place else ""),
        }
    )
    return output


def ensure_input_columns(fieldnames: list[str] | None) -> list[str]:
    if not fieldnames:
        raise ValueError("Input CSV is empty or missing headers.")
    if "address" not in fieldnames:
        raise ValueError("Input CSV must include an 'address' column.")
    return fieldnames


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as source:
        reader = csv.DictReader(source)
        fieldnames = ensure_input_columns(reader.fieldnames)
        return fieldnames, [dict(row) for row in reader]


def openpyxl_or_exit():
    try:
        import openpyxl  # type: ignore
    except ImportError as error:
        raise RuntimeError(
            "Excel files require openpyxl. Run this script with the bundled "
            "Python runtime shown in README.md, or install openpyxl for your Python."
        ) from error
    return openpyxl


def read_xlsx_rows(path: Path, sheet_name: str | None) -> tuple[list[str], list[dict[str, str]]]:
    openpyxl = openpyxl_or_exit()
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Input workbook sheet is empty.")

    headers = [compact(value) for value in rows[0]]
    while headers and not headers[-1]:
        headers.pop()
    fieldnames = ensure_input_columns(headers)

    records: list[dict[str, str]] = []
    for values in rows[1:]:
        record = {
            header: compact(values[index] if index < len(values) else "")
            for index, header in enumerate(fieldnames)
        }
        if any(record.values()):
            records.append(record)
    return fieldnames, records


def read_rows(path: Path, sheet_name: str | None) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix == ".xlsx":
        return read_xlsx_rows(path, sheet_name)
    raise ValueError("Input must be a .csv or .xlsx file.")


def write_csv_rows(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as destination:
        writer = csv.DictWriter(destination, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx_rows(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    openpyxl = openpyxl_or_exit()
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Enriched Addresses"
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    if rows:
        table_ref = f"A1:{sheet.cell(row=len(rows) + 1, column=len(fieldnames)).coordinate}"
        table = Table(displayName="EnrichedAddresses", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        header = compact(column_cells[0].value)
        max_len = max(len(compact(cell.value)) for cell in column_cells[: min(len(rows) + 1, 100)])
        width = min(max(max_len + 2, len(header) + 2, 10), 48)
        sheet.column_dimensions[column_cells[0].column_letter].width = width

    workbook.save(path)


def write_rows(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        write_csv_rows(path, fieldnames, rows)
        return
    if suffix == ".xlsx":
        write_xlsx_rows(path, fieldnames, rows)
        return
    raise ValueError("Output must be a .csv or .xlsx file.")


def enrich_file(
    input_path: Path,
    output_path: Path,
    provider: str = "auto",
    cache_path: Path = Path(".cache/places_cache.json"),
    limit: int | None = None,
    sheet: str | None = None,
    max_results: int = 1,
    search_strategy: str = "exact",
    sleep_seconds: float = 0.1,
    field_preset: str = "contact",
    field_mask: str | None = None,
) -> tuple[int, int]:
    load_env_file(Path(__file__).with_name(".env"))
    api_key = os.environ.get("GOOGLE_PLACES_API_KEY")

    if provider == "auto":
        provider = "google" if api_key else "heuristic"
    if provider == "google" and not api_key:
        raise RuntimeError("Provider google requires GOOGLE_PLACES_API_KEY in .env or the environment.")

    resolved_field_mask = field_mask or FIELD_PRESETS[field_preset]
    cache = load_cache(cache_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    input_columns, input_rows = read_rows(input_path, sheet)
    fieldnames = input_columns + [column for column in OUTPUT_COLUMNS if column not in input_columns]
    if "match_rank" not in fieldnames:
        fieldnames.append("match_rank")

    output_rows: list[dict[str, str]] = []
    processed = 0
    for row in input_rows:
        if limit is not None and processed >= limit:
            break
        places: list[dict[str, Any]] = []
        status = "no_match"
        error = ""
        seen_place_ids: set[str] = set()
        for query in build_search_queries(row, search_strategy):
            query_places, query_status, query_error = lookup_places(
                query=query,
                provider=provider,
                api_key=api_key,
                field_mask=resolved_field_mask,
                cache=cache,
                sleep_seconds=sleep_seconds,
                max_results=max_results,
            )
            if query_error and not error:
                error = query_error
            if query_status == "api_error":
                status = "api_error"
                continue
            if query_places:
                status = "enriched"
            elif status != "enriched":
                status = query_status
            for place in query_places:
                place_id = compact(place.get("id")) or json.dumps(place, sort_keys=True)
                if place_id in seen_place_ids:
                    continue
                seen_place_ids.add(place_id)
                place_with_query = dict(place)
                place_with_query["_query_used"] = query
                places.append(place_with_query)
                if len(places) >= max_results:
                    break
            if len(places) >= max_results or provider == "heuristic":
                break
        if places:
            for rank, place in enumerate(places, start=1):
                output_rows.append(enriched_row(row, place, provider, status, error, rank))
        else:
            output_rows.append(enriched_row(row, None, provider, status, error))
        processed += 1

    write_rows(output_path, fieldnames, output_rows)

    save_cache(cache_path, cache)
    return processed, len(output_rows)


def run() -> int:
    args = parse_args()
    try:
        processed, output_count = enrich_file(
            input_path=Path(args.input),
            output_path=Path(args.output),
            provider=args.provider,
            cache_path=Path(args.cache),
            limit=args.limit,
            sheet=args.sheet,
            max_results=args.max_results,
            search_strategy=args.search_strategy,
            sleep_seconds=args.sleep,
            field_preset=args.field_preset,
            field_mask=args.field_mask,
        )
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 2

    provider = args.provider
    if provider == "auto":
        provider = "google" if os.environ.get("GOOGLE_PLACES_API_KEY") else "heuristic"
    print(f"Processed {processed} rows with provider={provider}.")
    print(f"Wrote {output_count} output rows.")
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
